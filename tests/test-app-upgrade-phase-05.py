#!/usr/bin/env python3
import os
import sys
import subprocess
from PIL import Image
import openpyxl

def log_success(msg):
    print(f"[\033[92mSUCCESS\033[0m] {msg}")

def log_info(msg):
    print(f"[\033[94mINFO\033[0m] {msg}")

def log_failure(msg):
    print(f"[\033[91mFAILURE\033[0m] {msg}")
    sys.exit(1)

def get_flutter_cmd():
    flutter_cmd = "flutter"
    home = os.path.expanduser("~")
    local_paths = [
        os.path.join(home, "development", "flutter", "bin", "flutter"),
        os.path.join(home, "flutter", "bin", "flutter"),
        "/opt/flutter/bin/flutter"
    ]
    for p in local_paths:
        if os.path.exists(p):
            return p
    return flutter_cmd

def verify_phase_01_excel_formatting(workspace_dir):
    log_info("=== [Phase 01] Verifying Excel Export Formatting & Auto-Fit ===")
    service_file = os.path.join(workspace_dir, "lib", "services", "excel_service.dart")
    if not os.path.exists(service_file):
        log_failure(f"File not found: {service_file}")

    with open(service_file, "r", encoding="utf-8") as f:
        content = f.read()

    if "'Times New Roman'" not in content:
        log_failure("ExcelService does not configure 'Times New Roman' font!")
    log_success("Verified 'Times New Roman' font family configured in ExcelService.")

    if "fontSize: 12" not in content:
        log_failure("ExcelService does not configure fontSize: 12!")
    log_success("Verified fontSize: 12 configured.")

    if "sheet.setRowHeight(0, 28.0)" not in content:
        log_failure("Header row height 28.0 not found in ExcelService!")
    log_success("Verified header row height: 28.0.")

    if "sheet.setRowHeight(rowIdx, 24.0)" not in content:
        log_failure("Data row height 24.0 not found in ExcelService!")
    log_success("Verified data row height: 24.0.")

    if "12.0" not in content or "5.0" not in content:
        log_failure("Auto-fit metrics (minimum 12.0 or padding 5.0) not found!")
    log_success("Verified column auto-fit sizing logic.")

    # Deep verify generated excel file if available
    excel_test_file = os.path.join(workspace_dir, "test", "test_out.xlsx")
    if os.path.exists(excel_test_file):
        wb = openpyxl.load_workbook(excel_test_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value is not None and cell.font:
                    if cell.font.name and "Times New Roman" not in cell.font.name:
                        log_failure(f"Cell {cell.coordinate} font is {cell.font.name}, expected Times New Roman")
                    if cell.font.size and cell.font.size != 12:
                        log_failure(f"Cell {cell.coordinate} font size is {cell.font.size}, expected 12")
        log_success("Verified generated Excel sheet formatting properties via openpyxl.")

def verify_phase_02_drug_sorting(workspace_dir):
    log_info("=== [Phase 02] Verifying Drug List Alphabetical Sorting (A-Z) ===")
    # 1. DrugItem
    drug_item_file = os.path.join(workspace_dir, "lib", "models", "drug_item.dart")
    with open(drug_item_file, "r", encoding="utf-8") as f:
        drug_item_src = f.read()

    if "compareByName" not in drug_item_src:
        log_failure("compareByName not found in drug_item.dart!")
    if "sortAndReindex" not in drug_item_src:
        log_failure("sortAndReindex not found in drug_item.dart!")
    log_success("Verified DrugItem compareByName and sortAndReindex methods.")

    # 2. DashboardPage
    dashboard_file = os.path.join(workspace_dir, "lib", "views", "dashboard_page.dart")
    with open(dashboard_file, "r", encoding="utf-8") as f:
        dashboard_src = f.read()

    if "DrugItem.sortAndReindex(_items)" not in dashboard_src:
        log_failure("DashboardPage _reindexItems does not call DrugItem.sortAndReindex!")
    log_success("Verified DashboardPage auto-sorts on reindex.")

    # 3. ExcelService
    excel_file = os.path.join(workspace_dir, "lib", "services", "excel_service.dart")
    with open(excel_file, "r", encoding="utf-8") as f:
        excel_src = f.read()

    if "DrugItem.sortAndReindex(items)" not in excel_src:
        log_failure("ExcelService generateExcel does not sort items before export!")
    log_success("Verified ExcelService auto-sorts before export.")

def verify_phase_03_app_renaming(workspace_dir):
    log_info("=== [Phase 03] Verifying Application Renaming to 'Tạo bill thuốc' ===")
    target_name = "Tạo bill thuốc"

    # 1. lib/main.dart
    main_dart = os.path.join(workspace_dir, "lib", "main.dart")
    with open(main_dart, "r", encoding="utf-8") as f:
        src = f.read()
    if f"title: '{target_name}'" not in src:
        log_failure(f"lib/main.dart does not contain title: '{target_name}'")
    log_success("Verified lib/main.dart title.")

    # 2. lib/views/dashboard_page.dart
    dashboard_page = os.path.join(workspace_dir, "lib", "views", "dashboard_page.dart")
    with open(dashboard_page, "r", encoding="utf-8") as f:
        src = f.read()
    if f"'{target_name}'" not in src:
        log_failure(f"lib/views/dashboard_page.dart does not contain '{target_name}'")
    log_success("Verified lib/views/dashboard_page.dart title.")

    # 3. android/app/src/main/AndroidManifest.xml
    android_manifest = os.path.join(workspace_dir, "android", "app", "src", "main", "AndroidManifest.xml")
    with open(android_manifest, "r", encoding="utf-8") as f:
        src = f.read()
    if f'android:label="{target_name}"' not in src:
        log_failure(f"AndroidManifest.xml does not contain android:label=\"{target_name}\"")
    log_success("Verified AndroidManifest.xml label.")

    # 4. linux/runner/my_application.cc
    linux_app = os.path.join(workspace_dir, "linux", "runner", "my_application.cc")
    with open(linux_app, "r", encoding="utf-8") as f:
        src = f.read()
    if f'"{target_name}"' not in src:
        log_failure(f"linux/runner/my_application.cc does not contain \"{target_name}\"")
    log_success("Verified linux/runner/my_application.cc window and header bar title.")

    # 5. windows/runner/main.cpp
    win_main = os.path.join(workspace_dir, "windows", "runner", "main.cpp")
    with open(win_main, "r", encoding="utf-8") as f:
        src = f.read()
    if f'L"{target_name}"' not in src:
        log_failure(f"windows/runner/main.cpp does not contain L\"{target_name}\"")
    log_success("Verified windows/runner/main.cpp window title.")

    # 6. windows/runner/Runner.rc
    win_rc = os.path.join(workspace_dir, "windows", "runner", "Runner.rc")
    with open(win_rc, "r", encoding="utf-8") as f:
        src = f.read()
    if f'VALUE "FileDescription", "{target_name}"' not in src or f'VALUE "ProductName", "{target_name}"' not in src:
        log_failure(f"windows/runner/Runner.rc does not contain \"{target_name}\" in FileDescription/ProductName")
    log_success("Verified windows/runner/Runner.rc metadata strings.")

    # 7. build-deb.sh
    build_deb = os.path.join(workspace_dir, "build-deb.sh")
    with open(build_deb, "r", encoding="utf-8") as f:
        src = f.read()
    if f"Name={target_name}" not in src:
        log_failure(f"build-deb.sh does not contain Name={target_name}")
    log_success("Verified build-deb.sh packaging metadata.")

    # 8. installer.iss & installer.nsi
    iss_file = os.path.join(workspace_dir, "installer.iss")
    with open(iss_file, "r", encoding="utf-8") as f:
        src = f.read()
    if f'#define AppName "{target_name}"' not in src:
        log_failure(f"installer.iss does not define AppName as \"{target_name}\"")
    log_success("Verified installer.iss AppName.")

    nsi_file = os.path.join(workspace_dir, "installer.nsi")
    with open(nsi_file, "r", encoding="utf-8") as f:
        src = f.read()
    if f'!define APP_NAME      "{target_name}"' not in src:
        log_failure(f"installer.nsi does not define APP_NAME as \"{target_name}\"")
    log_success("Verified installer.nsi APP_NAME.")

def verify_phase_04_adaptive_icons(workspace_dir):
    log_info("=== [Phase 04] Verifying Cross-Platform Adaptive App Icons ===")
    
    # 1. Master assets
    assets = {
        "app_icon.png": (1024, 1024),
        "app_icon_foreground.png": (1024, 1024),
        "app_icon_background.png": (1024, 1024),
    }
    for name, expected_size in assets.items():
        path = os.path.join(workspace_dir, "assets", "icon", name)
        if not os.path.isfile(path):
            log_failure(f"Missing master asset: {path}")
        with Image.open(path) as img:
            if img.size != expected_size or img.mode != "RGBA":
                log_failure(f"{name} invalid properties: size={img.size}, mode={img.mode}")
        log_success(f"Verified master asset: {name} ({expected_size[0]}x{expected_size[1]} RGBA)")

    # 2. Android resources
    res_dir = os.path.join(workspace_dir, "android", "app", "src", "main", "res")
    xml_files = [
        os.path.join(res_dir, "mipmap-anydpi-v26", "ic_launcher.xml"),
        os.path.join(res_dir, "mipmap-anydpi-v26", "ic_launcher_round.xml"),
        os.path.join(res_dir, "drawable", "ic_launcher_background.xml"),
    ]
    for xml in xml_files:
        if not os.path.isfile(xml):
            log_failure(f"Missing Android XML resource: {xml}")
    log_success("Verified Android adaptive XML definitions.")

    legacy_specs = {
        "mipmap-mdpi": (48, 48),
        "mipmap-hdpi": (72, 72),
        "mipmap-xhdpi": (96, 96),
        "mipmap-xxhdpi": (144, 144),
        "mipmap-xxxhdpi": (192, 192),
    }
    fg_specs = {
        "mipmap-mdpi": (108, 108),
        "mipmap-hdpi": (162, 162),
        "mipmap-xhdpi": (216, 216),
        "mipmap-xxhdpi": (324, 324),
        "mipmap-xxxhdpi": (432, 432),
    }
    for folder, size in legacy_specs.items():
        for ftype, s in [("ic_launcher.png", size), ("ic_launcher_round.png", size), ("ic_launcher_foreground.png", fg_specs[folder])]:
            p = os.path.join(res_dir, folder, ftype)
            if not os.path.isfile(p):
                log_failure(f"Missing {p}")
            with Image.open(p) as img:
                if img.size != s:
                    log_failure(f"{folder}/{ftype} size {img.size} != {s}")
    log_success("Verified all Android mipmap densities (mdpi to xxxhdpi).")

    # 3. Windows ICO
    ico_path = os.path.join(workspace_dir, "windows", "runner", "resources", "app_icon.ico")
    if not os.path.isfile(ico_path) or os.path.getsize(ico_path) < 10000:
        log_failure(f"Windows app_icon.ico is invalid or missing at {ico_path}")
    log_success("Verified Windows app_icon.ico bundle.")

    # 4. Linux icons & runner
    linux_icon = os.path.join(workspace_dir, "linux", "runner", "resources", "app_icon.png")
    if not os.path.isfile(linux_icon):
        log_failure(f"Missing Linux app_icon.png at {linux_icon}")
    with Image.open(linux_icon) as img:
        if img.size != (512, 512):
            log_failure(f"Linux app_icon.png size {img.size} != (512, 512)")
    my_app_cc = os.path.join(workspace_dir, "linux", "runner", "my_application.cc")
    with open(my_app_cc, "r", encoding="utf-8") as f:
        if "gtk_window_set_icon_from_file" not in f.read():
            log_failure("my_application.cc does not call gtk_window_set_icon_from_file")
    log_success("Verified Linux app icon and my_application.cc integration.")

def run_flutter_test_suite(workspace_dir):
    log_info("=== [Phase 05] Executing Full Flutter Test Suite ===")
    flutter_cmd = get_flutter_cmd()
    try:
        res = subprocess.run(
            [flutter_cmd, "test"],
            cwd=workspace_dir,
            capture_output=True,
            text=True,
            check=True
        )
        print(res.stdout)
        log_success("Full Flutter test suite passed cleanly!")
    except subprocess.CalledProcessError as e:
        print(e.stdout)
        print(e.stderr)
        log_failure(f"Flutter test suite failed: {e}")

def run_flutter_analyze(workspace_dir):
    log_info("=== [Phase 05] Executing Flutter Analyze (Zero Warnings Check) ===")
    flutter_cmd = get_flutter_cmd()
    try:
        res = subprocess.run(
            [flutter_cmd, "analyze"],
            cwd=workspace_dir,
            capture_output=True,
            text=True,
            check=True
        )
        print(res.stdout)
        if "No issues found!" in res.stdout or "0 issues found" in res.stdout:
            log_success("Flutter analyze completed with 0 issues / 0 warnings!")
        else:
            log_failure(f"Analyzer found issues:\n{res.stdout}")
    except subprocess.CalledProcessError as e:
        print(e.stdout)
        print(e.stderr)
        log_failure(f"Flutter analyze failed: {e}")

def main():
    print("=" * 70)
    print("RUNNING MASTER INTEGRATION & VERIFICATION SUITE: PHASE 05")
    print("=" * 70)

    workspace_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

    verify_phase_01_excel_formatting(workspace_dir)
    print("-" * 70)

    verify_phase_02_drug_sorting(workspace_dir)
    print("-" * 70)

    verify_phase_03_app_renaming(workspace_dir)
    print("-" * 70)

    verify_phase_04_adaptive_icons(workspace_dir)
    print("-" * 70)

    run_flutter_test_suite(workspace_dir)
    print("-" * 70)

    run_flutter_analyze(workspace_dir)
    print("-" * 70)

    print("=" * 70)
    log_success("ALL PHASE 01-05 END-TO-END INTEGRATION & REGRESSION CHECKS PASSED!")
    print("=" * 70)

if __name__ == "__main__":
    main()
