#!/usr/bin/env python3
import os
import sys
import subprocess
import openpyxl

def log_success(msg):
    print(f"[\033[92mSUCCESS\033[0m] {msg}")

def log_info(msg):
    print(f"[\033[94mINFO\033[0m] {msg}")

def log_failure(msg):
    print(f"[\033[91mFAILURE\033[0m] {msg}")
    sys.exit(1)

def find_flutter():
    log_info("Locating Flutter SDK...")
    # Try global path first
    try:
        res = subprocess.run(["flutter", "--version"], capture_output=True, text=True, check=True)
        version_output = res.stdout.strip()
        log_success(f"Flutter is available globally: {version_output.splitlines()[0]}")
        return "flutter"
    except Exception:
        pass
        
    # Check common local installation directories
    home = os.path.expanduser("~")
    local_paths = [
        os.path.join(home, "development", "flutter", "bin", "flutter"),
        os.path.join(home, "flutter", "bin", "flutter"),
        "/opt/flutter/bin/flutter"
    ]
    
    for path in local_paths:
        if os.path.exists(path):
            try:
                res = subprocess.run([path, "--version"], capture_output=True, text=True, check=True)
                version_output = res.stdout.strip()
                log_success(f"Flutter found at local path: {path}")
                log_info(f"Version: {version_output.splitlines()[0]}")
                return path
            except Exception:
                pass
                
    log_failure("Flutter SDK could not be found.")

def verify_build_compiles(workspace_dir, flutter_cmd):
    log_info("Compiling Flutter release build for Linux desktop...")
    try:
        res = subprocess.run(
            [flutter_cmd, "build", "linux", "--release"],
            cwd=workspace_dir,
            capture_output=True,
            text=True,
            check=True
        )
        print(res.stdout)
        log_success("Flutter release build completed successfully.")
    except subprocess.CalledProcessError as e:
        print(e.stdout)
        print(e.stderr)
        log_failure(f"Flutter build failed with error: {e}")

def verify_deb_packaging(workspace_dir):
    log_info("Executing Debian packaging script...")
    packaging_script = os.path.join(workspace_dir, "build-deb.sh")
    
    if not os.path.exists(packaging_script):
        log_failure(f"Missing packaging script: {packaging_script}")
        
    # Clean up previous deb package if any
    deb_packages = [f for f in os.listdir(workspace_dir) if f.endswith(".deb")]
    for db in deb_packages:
        try:
            os.remove(os.path.join(workspace_dir, db))
        except Exception:
            pass

    try:
        res = subprocess.run(
            ["bash", "build-deb.sh"],
            cwd=workspace_dir,
            capture_output=True,
            text=True,
            check=True
        )
        print(res.stdout)
        log_success("Debian packaging executed successfully.")
        
        # Check if deb package exists
        deb_packages = [f for f in os.listdir(workspace_dir) if f.endswith(".deb")]
        if not deb_packages:
            log_failure("No .deb installer package generated at repository root!")
            
        log_success(f"Generated Debian package: {deb_packages[0]}")
    except subprocess.CalledProcessError as e:
        print(e.stdout)
        print(e.stderr)
        log_failure(f"Debian packaging failed: {e}")

def run_excel_integration_test(workspace_dir, flutter_cmd):
    log_info("Running manual input Excel generation integration test...")
    try:
        res = subprocess.run(
            [flutter_cmd, "test", "test/manual_input_excel_test.dart"],
            cwd=workspace_dir,
            capture_output=True,
            text=True,
            check=True
        )
        print(res.stdout)
        log_success("Generated manual_test_output.xlsx successfully via Integration Test.")
    except subprocess.CalledProcessError as e:
        print(e.stdout)
        print(e.stderr)
        log_failure(f"Excel generation Integration Test failed: {e}")

def verify_excel_styles_and_content(file_path):
    log_info(f"Starting styling and content verification on: {file_path}")
    if not os.path.exists(file_path):
        log_failure(f"Excel output file not found at: {file_path}!")
        
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if "Sheet1" not in wb.sheetnames:
            log_failure("Missing sheet 'Sheet1' in Excel workbook!")
        log_success("Sheet 'Sheet1' exists.")
        
        sheet = wb["Sheet1"]
        
        # 1. Verify Headers
        expected_headers = ["STT", "Tên thuốc", "Thương hiệu", "Quy cách", "Số lượng"]
        actual_headers = [sheet.cell(1, col).value for col in range(1, 6)]
        for i, (expected, actual) in enumerate(zip(expected_headers, actual_headers)):
            if expected != actual:
                log_failure(f"Header mismatch at column {i+1}! Expected: '{expected}', Got: '{actual}'")
        log_success(f"Headers verified successfully: {actual_headers}")
        
        # 2. Verify Header formatting
        header_cell = sheet.cell(1, 1)
        if not header_cell.font or not header_cell.font.bold:
            log_failure("Header font style is not bold!")
        if header_cell.font.name != "Inter":
            log_failure(f"Header font name is '{header_cell.font.name}', expected 'Inter'!")
        if header_cell.font.size != 11:
            log_failure(f"Header font size is {header_cell.font.size}, expected 11!")
        log_success("Header font (Inter 11pt Bold) verified.")

        if not header_cell.fill or header_cell.fill.fill_type != "solid":
            log_failure("Header fill type is not solid!")
        color = header_cell.fill.start_color.rgb
        # Allow variations of slate gray (like E2E8F0 or FFE2E8F0 due to alpha channel)
        if not color or ("E2E8F0" not in color.upper()):
            log_failure(f"Header color '{color}' does not match Slate-gray '#E2E8F0'!")
        log_success(f"Header Slate-gray fill color verified: {color}")
        
        # 3. Verify manual entries contents and formatting
        # Row 2 check
        r2_stt = sheet.cell(2, 1).value
        r2_name = sheet.cell(2, 2).value
        r2_brand = sheet.cell(2, 3).value
        r2_quy_cach = sheet.cell(2, 4).value
        r2_qty = sheet.cell(2, 5).value
        
        if r2_stt != 1 or r2_name != 'Manual Drug A' or r2_brand != 'Brand X' or r2_quy_cach != 'Hộp 30 viên' or r2_qty != 5:
            log_failure(f"Row 2 data mismatch! Got: STT={r2_stt}, Name={r2_name}, Brand={r2_brand}, QuyCach={r2_quy_cach}, Qty={r2_qty}")
        log_success("Row 2 manual entry data verified.")

        # Row 3 check
        r3_stt = sheet.cell(3, 1).value
        r3_name = sheet.cell(3, 2).value
        r3_brand = sheet.cell(3, 3).value
        r3_quy_cach = sheet.cell(3, 4).value
        r3_qty = sheet.cell(3, 5).value
        
        if r3_stt != 2 or r3_name != 'Manual Drug B' or r3_brand != 'N/A' or r3_quy_cach != 'Chai 100ml' or r3_qty != 2:
            log_failure(f"Row 3 data mismatch! Got: STT={r3_stt}, Name={r3_name}, Brand={r3_brand}, QuyCach={r3_quy_cach}, Qty={r3_qty}")
        log_success("Row 3 manual entry data verified.")

        # 4. Verify alignments
        # STT centered (Row 2, Column 1)
        if not sheet.cell(2, 1).alignment or sheet.cell(2, 1).alignment.horizontal != "center":
            log_failure("STT column (col 1) alignment is not centered!")
        # Name left aligned (Row 2, Column 2)
        if sheet.cell(2, 2).alignment and sheet.cell(2, 2).alignment.horizontal not in [None, "left"]:
            log_failure("Name column (col 2) alignment is not left-aligned!")
        # Brand left aligned (Row 2, Column 3)
        if sheet.cell(2, 3).alignment and sheet.cell(2, 3).alignment.horizontal not in [None, "left"]:
            log_failure("Brand column (col 3) alignment is not left-aligned!")
        # Quy cach left aligned (Row 2, Column 4)
        if sheet.cell(2, 4).alignment and sheet.cell(2, 4).alignment.horizontal not in [None, "left"]:
            log_failure("Quy cach column (col 4) alignment is not left-aligned!")
        log_success("Data column alignments (STT centered, Name/Brand/QuyCach left-aligned) verified.")

        # 5. Verify auto-fit column widths
        for col_letter in ['B', 'C', 'D']:
            width = sheet.column_dimensions[col_letter].width
            if not width or width <= 10:
                log_failure(f"Column {col_letter} auto-fit width is too small: {width}!")
            log_success(f"Column {col_letter} width auto-fitted: {width}")
            
        print("=" * 60)
        log_success("EXCEL WORKBOOK STYLE VERIFICATION PASSED SUCCESSFULLY!")
        print("=" * 60)
        
        # Clean up output file
        os.remove(file_path)
        log_info("Cleaned up manual_test_output.xlsx.")
        
    except Exception as e:
        log_failure(f"Failed to verify excel spreadsheet: {e}")

def main():
    print("=" * 60)
    print("RUNNING AUTOMATED MANUAL DRUG INPUT PHASE 03 INTEGRATION & E2E SUITE")
    print("=" * 60)
    
    workspace_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    flutter_cmd = find_flutter()
    
    # 1. Compile release executable
    verify_build_compiles(workspace_dir, flutter_cmd)
    print("-" * 60)
    
    # 2. Package release package (.deb)
    verify_deb_packaging(workspace_dir)
    print("-" * 60)
    
    # 3. Verify excel integration output
    run_excel_integration_test(workspace_dir, flutter_cmd)
    excel_file = os.path.join(workspace_dir, "manual_test_output.xlsx")
    verify_excel_styles_and_content(excel_file)
    print("-" * 60)
    
    log_success("ALL PHASE 03 E2E VERIFICATION AND PACKAGING CHECKS PASSED!")
    print("=" * 60)

if __name__ == "__main__":
    main()
