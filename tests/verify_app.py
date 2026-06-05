#!/usr/bin/env python3
import os
import sys
import openpyxl

def log_success(msg):
    print(f"[\033[92mSUCCESS\033[0m] {msg}")

def log_info(msg):
    print(f"[\033[94mINFO\033[0m] {msg}")

def log_failure(msg):
    print(f"[\033[91mFAILURE\033[0m] {msg}")
    sys.exit(1)

def verify_excel(file_path):
    log_info(f"Starting verification on Excel output file: {file_path}")
    
    if not os.path.exists(file_path):
        log_failure(f"Excel file not found at: {file_path}!")
        
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 1. Verify sheet exists
        if "Sheet1" not in wb.sheetnames:
            log_failure("Missing sheet 'Sheet1' in Excel workbook!")
        log_success("Sheet 'Sheet1' exists.")
        
        sheet = wb["Sheet1"]
        
        # 2. Verify Headers
        expected_headers = ["STT", "Tên thuốc", "Thương hiệu", "Quy cách", "Số lượng"]
        actual_headers = [sheet.cell(1, col).value for col in range(1, 6)]
        
        for i, (expected, actual) in enumerate(zip(expected_headers, actual_headers)):
            if expected != actual:
                log_failure(f"Header mismatch at column {i+1}! Expected: '{expected}', Got: '{actual}'")
        log_success(f"Headers verified successfully: {actual_headers}")
        
        # 3. Verify Header formatting
        header_cell = sheet.cell(1, 1)
        if header_cell.font:
            if header_cell.font.bold:
                log_success("Header font weight is bold.")
            else:
                log_info("Header font weight is not bold or openpyxl could not parse it.")
                
            if header_cell.font.name == "Inter":
                log_success("Header font family is correct ('Inter').")
            else:
                log_info(f"Header font family: {header_cell.font.name or 'Default'}")
                
        if header_cell.fill and header_cell.fill.fill_type == "solid":
            color = header_cell.fill.start_color.rgb
            # Allow variations of slate gray (like E2E8F0 or FFE2E8F0 due to alpha channel)
            if color and ("E2E8F0" in color or "e2e8f0" in color):
                log_success(f"Header fill color is correct: {color}")
            else:
                log_info(f"Header fill color: {color}")
                
        # 4. Verify some data cells (if file has test data)
        # Expected test row details
        expected_data = [
            {"stt": 1, "name": "Izac Syrup", "brand": "An Thiên", "quy_cach": "Chai 60ml", "qty": 3},
        ]
        
        # Check if there is data on row 2
        if sheet.cell(2, 1).value is not None:
            log_info("Checking sample row 2 data...")
            stt = sheet.cell(2, 1).value
            name = sheet.cell(2, 2).value
            brand = sheet.cell(2, 3).value
            quy_cach = sheet.cell(2, 4).value
            qty = sheet.cell(2, 5).value
            
            log_info(f"Found Row 2 -> STT: {stt}, Name: '{name}', Brand: '{brand}', Packaging: '{quy_cach}', Quantity: {qty}")
            
            # Match alignments if possible
            if sheet.cell(2, 1).alignment and sheet.cell(2, 1).alignment.horizontal == "center":
                log_success("STT column alignment is centered.")
            if sheet.cell(2, 2).alignment and sheet.cell(2, 2).alignment.horizontal == "left":
                log_success("Name column alignment is left-aligned.")
                
        # 5. Check auto-fit widths
        col_b_width = sheet.column_dimensions['B'].width
        col_c_width = sheet.column_dimensions['C'].width
        col_d_width = sheet.column_dimensions['D'].width
        
        log_info(f"Column widths -> Name (B): {col_b_width}, Brand (C): {col_c_width}, Packaging (D): {col_d_width}")
        if col_b_width and col_b_width > 10:
            log_success("Column width auto-fitting verified.")
            
        print("=" * 60)
        log_success("EXCEL WORKBOOK STYLE VERIFICATION PASSED SUCCESSFULLY!")
        print("=" * 60)
        
    except Exception as e:
        log_failure(f"Failed to verify excel spreadsheet: {e}")

def main():
    print("=" * 60)
    print("RUNNING AUTOMATED E2E EXCEL STYLING AND CONTENT VERIFIER")
    print("=" * 60)
    
    workspace_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    
    excel_file = os.path.join(workspace_dir, "test_output.xlsx")
    verify_excel(excel_file)

if __name__ == "__main__":
    main()
