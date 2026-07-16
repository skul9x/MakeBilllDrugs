#!/usr/bin/env python3
import os
import sys
import subprocess
import openpyxl

def log_success(msg):
    print(f"[\033[92mSUCCESS\033[0m] {msg}")

def log_info(msg):
    print(f"[\033[94mINFO\033[0m] {msg}")

def log_failure(msg):
    print(f"[\033[91mFAILURE\033[0m] {msg}")
    sys.exit(1)

def find_flutter():
    log_info("Locating Flutter SDK...")
    try:
        res = subprocess.run(["flutter", "--version"], capture_output=True, text=True, check=True)
        version_output = res.stdout.strip()
        log_success(f"Flutter is available globally: {version_output.splitlines()[0]}")
        return "flutter"
    except Exception:
        pass
        
    home = os.path.expanduser("~")
    local_paths = [
        os.path.join(home, "development", "flutter", "bin", "flutter"),
        os.path.join(home, "flutter", "bin", "flutter"),
        "/opt/flutter/bin/flutter"
    ]
    for path in local_paths:
        if os.path.exists(path):
            try:
                res = subprocess.run([path, "--version"], capture_output=True, text=True, check=True)
                version_output = res.stdout.strip()
                log_success(f"Flutter found at local path: {path}")
                return path
            except Exception:
                pass
    log_failure("Flutter SDK not found.")

def main():
    print("=" * 60)
    print("RUNNING AUTOMATED E2E INTEGRATION TESTS FOR LONG CHAU PHASE 02")
    print("=" * 60)
    
    flutter_bin = find_flutter()
    workspace_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    
    # 1. Verify files exist
    required_files = [
        "lib/models/drug_info.dart",
        "lib/models/drug_item.dart",
        "lib/services/drug_parser.dart",
        "lib/services/excel_service.dart",
        "test/long_chau_integration_test.dart",
    ]
    for f in required_files:
        p = os.path.join(workspace_dir, f)
        if not os.path.exists(p):
            log_failure(f"Required file is missing: {f}")
        log_success(f"Verified existence of: {f}")
        
    # 2. Run Dart Integration Test
    log_info("Running Long Chau Dart integration tests...")
    try:
        res = subprocess.run(
            [flutter_bin, "test", "test/long_chau_integration_test.dart"],
            cwd=workspace_dir,
            capture_output=True,
            text=True
        )
        print(res.stdout)
        if res.returncode == 0:
            log_success("Long Chau Integration Dart tests passed successfully!")
        else:
            print(res.stderr)
            log_failure(f"Dart Integration tests failed with code {res.returncode}")
    except Exception as e:
        log_failure(f"Failed to execute integration tests: {e}")

    # 3. Verify Output Excel using openpyxl
    log_info("Verifying generated Excel spreadsheet structure and styling...")
    excel_path = os.path.join(workspace_dir, "test", "long_chau_test_out.xlsx")
    if not os.path.exists(excel_path):
        log_failure("Excel output file does not exist!")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "Sheet1" not in wb.sheetnames:
            log_failure("Sheet 'Sheet1' is missing in the generated Excel file!")
        log_success("Sheet 'Sheet1' exists in Excel workbook.")

        sheet = wb["Sheet1"]

        # Verify Headers
        expected_headers = ["STT", "Tên thuốc", "Thương hiệu", "Quy cách", "Số lượng"]
        actual_headers = [sheet.cell(1, col).value for col in range(1, 6)]
        for idx, (exp, act) in enumerate(zip(expected_headers, actual_headers)):
            if exp != act:
                log_failure(f"Header mismatch at column {idx+1}! Expected: '{exp}', Got: '{act}'")
        log_success("Column headers are correct: STT, Tên thuốc, Thương hiệu, Quy cách, Số lượng")

        # Verify Header Styles
        h_cell = sheet.cell(1, 1)
        if h_cell.font:
            if h_cell.font.bold:
                log_success("Header font is bold.")
            if h_cell.font.name == "Inter":
                log_success("Header font family is correct ('Inter').")
            if h_cell.font.size == 11:
                log_success("Header font size is 11.")

        if h_cell.fill and h_cell.fill.fill_type == "solid":
            color = h_cell.fill.start_color.rgb
            if color and ("E2E8F0" in color or "e2e8f0" in color):
                log_success(f"Header fill color is correct (#E2E8F0): {color}")

        # Verify Deduplicated content
        # Row 2: Paracetamol Stada (7 items)
        # Row 3: Augmentin (1 item)
        # Row 4: Hapacol 150 (3 items)
        expected_rows = [
            (1, "Viên nén Paracetamol Stada 500mg điều trị các cơn đau đầu, đau thần kinh, đau răng (10 vỉ x 10 viên)", "DHG", "Hộp 10 Vỉ x 10 Viên", 7),
            (2, "Thuốc Augmentin 1g GSK điều trị nhiễm khuẩn (2 vỉ x 7 viên)", "SMITHKLINE BEECHAM PHARMACEUTICALS", "Hộp 2 Vỉ x 7 Viên", 1),
            (3, "Bột Hapacol 150 DHG giảm đau, hạ sốt (24 gói)", "DHG", "Hộp 24 Gói", 3)
        ]

        for i, (exp_stt, exp_name, exp_brand, exp_quycach, exp_qty) in enumerate(expected_rows):
            r = i + 2
            act_stt = sheet.cell(r, 1).value
            act_name = sheet.cell(r, 2).value
            act_brand = sheet.cell(r, 3).value
            act_quycach = sheet.cell(r, 4).value
            act_qty = sheet.cell(r, 5).value

            if act_stt != exp_stt:
                log_failure(f"Row {r} STT mismatch: Expected {exp_stt}, got {act_stt}")
            if act_name != exp_name:
                log_failure(f"Row {r} Name mismatch: Expected '{exp_name}', got '{act_name}'")
            if act_brand != exp_brand:
                log_failure(f"Row {r} Brand mismatch: Expected '{exp_brand}', got '{act_brand}'")
            if act_quycach != exp_quycach:
                log_failure(f"Row {r} Quy cách mismatch: Expected '{exp_quycach}', got '{act_quycach}'")
            if act_qty != exp_qty:
                log_failure(f"Row {r} Quantity mismatch: Expected {exp_qty}, got {act_qty}")

            log_success(f"Verified Row {r} (STT {act_stt}, Qty {act_qty}) content successfully.")

        # Check Column dimensions auto-fitting
        col_b_width = sheet.column_dimensions['B'].width
        col_c_width = sheet.column_dimensions['C'].width
        col_d_width = sheet.column_dimensions['D'].width
        if col_b_width and col_b_width > 10:
            log_success(f"Auto-fit Column Widths verified: Name={col_b_width}, Brand={col_c_width}, Packaging={col_d_width}")

    except Exception as e:
        log_failure(f"Excel verification failed with error: {e}")

    print("=" * 60)
    log_success("LONG CHAU PHASE 02 E2E VERIFICATION COMPLETED SUCCESSFULLY!")
    print("=" * 60)

if __name__ == "__main__":
    main()
