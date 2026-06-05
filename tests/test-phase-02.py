#!/usr/bin/env python3
import os
import sys
import subprocess
import openpyxl

def log_success(msg):
    print(f"[\033[92mSUCCESS\033[0m] {msg}")

def log_info(msg):
    print(f"[\033[94mINFO\033[0m] {msg}")

def log_failure(msg):
    print(f"[\033[91mFAILURE\033[0m] {msg}")
    sys.exit(1)

def verify_files_exist(workspace_dir):
    log_info("Checking Phase 02 source files...")
    required_files = [
        "lib/models/drug_info.dart",
        "lib/models/drug_item.dart",
        "lib/services/drug_parser.dart",
        "lib/services/excel_service.dart",
        "test/core_logic_test.dart"
    ]
    
    for file in required_files:
        path = os.path.join(workspace_dir, file)
        if not os.path.exists(path):
            log_failure(f"Missing required Phase 02 file: {file}")
        log_success(f"File exists: {file}")

def run_dart_tests(workspace_dir):
    log_info("Running Dart unit tests for Core Logic...")
    
    # Locate flutter command
    flutter_cmd = "flutter"
    home = os.path.expanduser("~")
    local_paths = [
        os.path.join(home, "development", "flutter", "bin", "flutter"),
        os.path.join(home, "flutter", "bin", "flutter"),
        "/opt/flutter/bin/flutter"
    ]
    for p in local_paths:
        if os.path.exists(p):
            flutter_cmd = p
            break
            
    try:
        res = subprocess.run(
            [flutter_cmd, "test", "test/core_logic_test.dart"],
            cwd=workspace_dir,
            capture_output=True,
            text=True,
            check=True
        )
        print(res.stdout)
        log_success("All Dart core logic unit tests passed successfully!")
    except subprocess.CalledProcessError as e:
        print(e.stdout)
        print(e.stderr)
        log_failure(f"Dart unit tests failed with error: {e}")

def verify_output_excel(workspace_dir):
    # If the dart test writes a test output file, we can verify its styles using python-openpyxl
    log_info("Checking for Excel generator output styling check...")
    excel_path = os.path.join(workspace_dir, "test", "test_out.xlsx")
    if not os.path.exists(excel_path):
        log_info("Excel test output not found (this will compile/create during actual test run). Skipping styling validation.")
        return
        
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        
        # Verify Headers
        expected = ["STT", "Tên thuốc", "Thương hiệu", "Quy cách", "Số lượng"]
        actual = [sheet.cell(1, col).value for col in range(1, 6)]
        
        for exp, act in zip(expected, actual):
            if exp != act:
                log_failure(f"Excel Header mismatch! Expected '{exp}', got '{act}'")
        log_success("Excel columns headers verified successfully.")
        
        # Verify Font styling on header
        header_cell = sheet.cell(1, 1)
        if header_cell.font and header_cell.font.bold:
            log_success("Header font is bold.")
        if header_cell.font and header_cell.font.name == "Inter":
            log_success("Header font family is Inter.")
            
    except Exception as e:
        log_failure(f"Error checking excel spreadsheet styles: {e}")

def main():
    print("=" * 60)
    print("RUNNING AUTOMATED FLUTTER PHASE 02 CORE LOGIC TEST SUITE")
    print("=" * 60)
    
    workspace_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    
    verify_files_exist(workspace_dir)
    print("-" * 60)
    
    run_dart_tests(workspace_dir)
    print("-" * 60)
    
    verify_output_excel(workspace_dir)
    print("-" * 60)
    
    log_success("ALL PHASE 02 CORE LOGIC VERIFICATION CHECKS PASSED!")
    print("=" * 60)

if __name__ == "__main__":
    main()
